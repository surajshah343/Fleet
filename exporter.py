"""Writing results back out to Excel.

The obvious approach -- ``openpyxl.load_workbook(path)``, add sheets, save --
does not survive these files.  Loading the 60 MB source workbook needs several
gigabytes of RAM and gets killed on a normal machine.

So :func:`append_sheets` works at the ZIP level instead.  An ``.xlsx`` is a ZIP
of XML parts; adding a worksheet means copying every existing part across
untouched and editing three small ones:

* ``[Content_Types].xml``      -- declare the new parts
* ``xl/workbook.xml``          -- list the new sheets
* ``xl/_rels/workbook.xml.rels`` -- point the listings at the parts

Memory stays flat because parts are streamed one at a time, and every original
sheet keeps its formatting, formulas and pivot tables byte for byte.

:func:`write_new_workbook` is the lightweight alternative: a brand new file
containing only the tabs we generated.
"""

from __future__ import annotations

import re
import shutil
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from xml.sax.saxutils import escape

import numpy as np
import pandas as pd

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
SHEET_CT = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
SHEET_REL_TYPE = f"{REL_NS}/worksheet"

MAX_CELL_CHARS = 32767
EXCEL_EPOCH = datetime(1899, 12, 30)
_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
INVALID_SHEET_CHARS = r"[]:*?/\\"


@dataclass
class SheetPayload:
    """One sheet to write."""

    name: str
    frame: pd.DataFrame


def safe_sheet_name(name: str, taken: Iterable[str] = ()) -> str:
    """Make a sheet name Excel will accept and that is not already used."""
    cleaned = "".join(" " if c in INVALID_SHEET_CHARS else c for c in str(name)).strip()
    cleaned = (cleaned or "Sheet")[:31]
    lowered = {t.lower() for t in taken}
    if cleaned.lower() not in lowered:
        return cleaned
    stem = cleaned[:28]
    for suffix in range(2, 100):
        candidate = f"{stem}_{suffix}"[:31]
        if candidate.lower() not in lowered:
            return candidate
    return cleaned


def column_letter(index: int) -> str:
    """1-based column index -> ``A``, ``B`` ... ``AA``."""
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


_LETTERS = [column_letter(i) for i in range(1, 1000)]


def _clean_text(value: str) -> str:
    text = _ILLEGAL_XML.sub("", str(value))
    if len(text) > MAX_CELL_CHARS:
        text = text[: MAX_CELL_CHARS - 1] + "…"
    return escape(text)


def _excel_serial(value) -> float:
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime(value.year, value.month, value.day)
    if isinstance(value, time):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400.0
    delta = value - EXCEL_EPOCH
    return delta.days + delta.seconds / 86400.0


def _cell_xml(ref: str, value, date_style: Optional[int]) -> str:
    """Render one ``<c>`` element."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    if value is pd.NaT or (isinstance(value, pd.Timestamp) and pd.isna(value)):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date, time)):
        style = f' s="{date_style}"' if date_style is not None else ""
        return f'<c r="{ref}"{style}><v>{_excel_serial(value):.10g}</v></c>'
    if isinstance(value, (bool, np.bool_)):
        return f'<c r="{ref}" t="b"><v>{int(value)}</v></c>'
    if isinstance(value, (int, np.integer)):
        return f'<c r="{ref}"><v>{int(value)}</v></c>'
    if isinstance(value, (float, np.floating)):
        if not np.isfinite(value):
            return ""
        return f'<c r="{ref}"><v>{float(value):.15g}</v></c>'
    text = _clean_text(value)
    if not text:
        return ""
    return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{text}</t></is></c>'


def _sheet_xml(
    frame: pd.DataFrame,
    header_style: Optional[int] = None,
    date_style: Optional[int] = None,
) -> Iterable[str]:
    """Yield the worksheet XML in chunks so nothing large is held in memory."""
    n_columns = max(frame.shape[1], 1)
    last_column = column_letter(n_columns)
    n_rows = len(frame) + 1

    yield (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{MAIN_NS}" xmlns:r="{REL_NS}">'
        f'<dimension ref="A1:{last_column}{max(n_rows, 1)}"/>'
        '<sheetViews><sheetView workbookViewId="0">'
        '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
        "</sheetView></sheetViews>"
        '<sheetFormatPr defaultRowHeight="15"/>'
        "<sheetData>"
    )

    style = f' s="{header_style}"' if header_style is not None else ""
    header_cells = "".join(
        f'<c r="{_LETTERS[i] if i < len(_LETTERS) else column_letter(i + 1)}1"'
        f'{style} t="inlineStr"><is><t xml:space="preserve">'
        f"{_clean_text(col)}</t></is></c>"
        for i, col in enumerate(frame.columns)
    )
    yield f'<row r="1">{header_cells}</row>'

    columns = [frame[c].to_numpy(dtype=object) for c in frame.columns]
    buffer: List[str] = []
    for row_index in range(len(frame)):
        excel_row = row_index + 2
        cells = []
        for col_index, column in enumerate(columns):
            letter = (
                _LETTERS[col_index]
                if col_index < len(_LETTERS)
                else column_letter(col_index + 1)
            )
            cell = _cell_xml(f"{letter}{excel_row}", column[row_index], date_style)
            if cell:
                cells.append(cell)
        buffer.append(f'<row r="{excel_row}">{"".join(cells)}</row>')
        if len(buffer) >= 2000:
            yield "".join(buffer)
            buffer.clear()
    if buffer:
        yield "".join(buffer)

    yield (
        "</sheetData>"
        f'<autoFilter ref="A1:{last_column}{max(n_rows, 1)}"/>'
        "</worksheet>"
    )


# --------------------------------------------------------------------------
# Styles: one bold header format and one date format
# --------------------------------------------------------------------------

def _extend_styles(styles_xml: bytes) -> Tuple[bytes, Optional[int], Optional[int]]:
    """Append a bold font and a date number format; return their style ids."""
    try:
        text = styles_xml.decode("utf-8")
        fonts_match = re.search(r"<fonts\b[^>]*>", text)
        xfs_match = re.search(r"<cellXfs\b[^>]*>", text)
        if not fonts_match or not xfs_match:
            return styles_xml, None, None

        font_count = len(re.findall(r"<font\b", text[: text.index("</fonts>")]))
        text = text.replace(
            "</fonts>", "<font><b/><sz val=\"11\"/><name val=\"Calibri\"/></font></fonts>", 1
        )
        text = re.sub(
            r'(<fonts\b[^>]*?)count="\d+"', rf'\1count="{font_count + 1}"', text, count=1
        )

        xfs_start = text.index("<cellXfs")
        xfs_end = text.index("</cellXfs>")
        xf_count = len(re.findall(r"<xf\b", text[xfs_start:xfs_end]))
        header_style = xf_count
        date_style = xf_count + 1
        new_xfs = (
            f'<xf numFmtId="0" fontId="{font_count}" fillId="0" borderId="0" '
            'xfId="0" applyFont="1"/>'
            '<xf numFmtId="14" fontId="0" fillId="0" borderId="0" xfId="0" '
            'applyNumberFormat="1"/>'
        )
        text = text.replace("</cellXfs>", new_xfs + "</cellXfs>", 1)
        text = re.sub(
            r'(<cellXfs\b[^>]*?)count="\d+"', rf'\1count="{xf_count + 2}"', text, count=1
        )
        return text.encode("utf-8"), header_style, date_style
    except Exception:  # noqa: BLE001 - styling is optional, never fail the export
        return styles_xml, None, None


# --------------------------------------------------------------------------
# Appending into an existing workbook
# --------------------------------------------------------------------------

def _existing_sheet_names(workbook_xml: bytes) -> List[str]:
    from xml.etree import ElementTree as ET

    root = ET.fromstring(workbook_xml)
    sheets = root.find(f"{{{MAIN_NS}}}sheets")
    if sheets is None:
        return []
    return [s.get("name", "") for s in sheets]


def append_sheets(
    source: str | Path,
    destination: str | Path,
    payloads: Sequence[SheetPayload],
    replace_existing: bool = True,
) -> List[str]:
    """Copy ``source`` to ``destination`` with ``payloads`` added as new sheets.

    Returns the sheet names actually written.  Existing sheets keep their
    formatting and formulas; only the workbook index is rewritten.
    """
    from xml.etree import ElementTree as ET

    source, destination = Path(source), Path(destination)
    ET.register_namespace("", MAIN_NS)
    ET.register_namespace("r", REL_NS)

    with zipfile.ZipFile(source) as archive:
        names = set(archive.namelist())
        workbook_xml = archive.read("xl/workbook.xml")
        rels_xml = archive.read("xl/_rels/workbook.xml.rels")
        content_types_xml = archive.read("[Content_Types].xml")
        styles_xml = archive.read("xl/styles.xml") if "xl/styles.xml" in names else None

    workbook = ET.fromstring(workbook_xml)
    sheets_element = workbook.find(f"{{{MAIN_NS}}}sheets")
    if sheets_element is None:
        raise ValueError("Workbook has no <sheets> element; file may be corrupt")

    rels_root = ET.fromstring(rels_xml)
    rel_targets: Dict[str, str] = {
        r.get("Id", ""): r.get("Target", "") for r in rels_root
    }

    # Work out which parts to drop when replacing same-named sheets.
    drop_parts: set[str] = set()
    wanted = {p.name.lower() for p in payloads}
    if replace_existing:
        for sheet in list(sheets_element):
            if sheet.get("name", "").lower() in wanted:
                rel_id = sheet.get(f"{{{REL_NS}}}id", "")
                target = rel_targets.get(rel_id, "")
                if target:
                    part = target if target.startswith("xl/") else f"xl/{target.lstrip('/')}"
                    drop_parts.add(part)
                sheets_element.remove(sheet)
                for rel in list(rels_root):
                    if rel.get("Id") == rel_id:
                        rels_root.remove(rel)

    # calcChain is a cache Excel rebuilds; dropping it avoids stale references.
    drop_parts.add("xl/calcChain.xml")
    for rel in list(rels_root):
        if rel.get("Target", "").endswith("calcChain.xml"):
            rels_root.remove(rel)

    taken_names = _existing_sheet_names(ET.tostring(workbook))
    used_rel_ids = {r.get("Id", "") for r in rels_root}
    used_sheet_ids = {
        int(s.get("sheetId", "0")) for s in sheets_element if s.get("sheetId", "0").isdigit()
    }
    used_parts = {n for n in names if n.startswith("xl/worksheets/sheet")}

    next_rel = max(
        [int(m.group(1)) for r in used_rel_ids if (m := re.match(r"rId(\d+)$", r))] or [0]
    )
    next_sheet_id = max(used_sheet_ids or {0})
    next_part = 0

    new_parts: List[Tuple[str, SheetPayload]] = []
    written_names: List[str] = []

    for payload in payloads:
        name = safe_sheet_name(payload.name, taken_names)
        taken_names.append(name)

        next_rel += 1
        rel_id = f"rId{next_rel}"
        while rel_id in used_rel_ids:
            next_rel += 1
            rel_id = f"rId{next_rel}"
        used_rel_ids.add(rel_id)

        next_sheet_id += 1
        next_part += 1
        part = f"xl/worksheets/sheet{next_part}.xml"
        while part in used_parts and part not in drop_parts:
            next_part += 1
            part = f"xl/worksheets/sheet{next_part}.xml"
        used_parts.add(part)

        sheet_element = ET.SubElement(sheets_element, f"{{{MAIN_NS}}}sheet")
        sheet_element.set("name", name)
        sheet_element.set("sheetId", str(next_sheet_id))
        sheet_element.set(f"{{{REL_NS}}}id", rel_id)

        rel = ET.SubElement(rels_root, f"{{{PKG_REL_NS}}}Relationship")
        rel.set("Id", rel_id)
        rel.set("Type", SHEET_REL_TYPE)
        rel.set("Target", part[len("xl/"):])

        new_parts.append((part, SheetPayload(name=name, frame=payload.frame)))
        written_names.append(name)

    # Content types
    ET.register_namespace("", CT_NS)
    content_types = ET.fromstring(content_types_xml)
    existing_overrides = {o.get("PartName") for o in content_types}
    for part, _ in new_parts:
        if f"/{part}" not in existing_overrides:
            override = ET.SubElement(content_types, f"{{{CT_NS}}}Override")
            override.set("PartName", f"/{part}")
            override.set("ContentType", SHEET_CT)
    for override in list(content_types):
        if override.get("PartName") == "/xl/calcChain.xml":
            content_types.remove(override)

    header_style = date_style = None
    new_styles = styles_xml
    if styles_xml is not None:
        new_styles, header_style, date_style = _extend_styles(styles_xml)

    ET.register_namespace("", MAIN_NS)
    workbook_out = ET.tostring(workbook, xml_declaration=True, encoding="UTF-8")
    ET.register_namespace("", PKG_REL_NS)
    rels_out = ET.tostring(rels_root, xml_declaration=True, encoding="UTF-8")
    ET.register_namespace("", CT_NS)
    content_types_out = ET.tostring(content_types, xml_declaration=True, encoding="UTF-8")

    rewritten = {
        "xl/workbook.xml": workbook_out,
        "xl/_rels/workbook.xml.rels": rels_out,
        "[Content_Types].xml": content_types_out,
    }
    if new_styles is not None:
        rewritten["xl/styles.xml"] = new_styles

    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(
        destination, "w", zipfile.ZIP_DEFLATED, compresslevel=6
    ) as out:
        for item in archive.infolist():
            if item.filename in drop_parts:
                continue
            if item.filename in rewritten:
                out.writestr(item.filename, rewritten.pop(item.filename))
                continue
            # Stream the part across without decompressing it all at once.
            with archive.open(item) as handle, out.open(item.filename, "w") as target:
                shutil.copyfileobj(handle, target, length=1 << 20)
        for name, data in rewritten.items():
            out.writestr(name, data)
        for part, payload in new_parts:
            with out.open(part, "w") as target:
                for chunk in _sheet_xml(payload.frame, header_style, date_style):
                    target.write(chunk.encode("utf-8"))

    return written_names


def write_new_workbook(
    destination: str | Path, payloads: Sequence[SheetPayload]
) -> List[str]:
    """Write a fresh workbook containing only ``payloads``.

    Uses openpyxl's write-only mode, which streams rows instead of building an
    in-memory tree.
    """
    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    written: List[str] = []
    for payload in payloads:
        name = safe_sheet_name(payload.name, written)
        sheet = workbook.create_sheet(title=name)
        sheet.append([str(c) for c in payload.frame.columns])
        for row in payload.frame.itertuples(index=False, name=None):
            sheet.append(
                [
                    None
                    if (isinstance(v, float) and not np.isfinite(v)) or v is pd.NaT
                    else v
                    for v in row
                ]
            )
        written.append(name)
    workbook.save(destination)
    return written
