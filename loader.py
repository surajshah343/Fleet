"""Reading the asset workbooks.

Two problems this module solves:

1. **The header is not on row 1.**  In the sample files the ``USA``, ``Canada``
   and ``US`` sheets put their header on row 2, while ``CAN`` puts it on row 4
   (rows above hold a title such as "Amounts in CAD$").  We sniff for it.
2. **Size.**  The workbooks are 50-60 MB.  Everything here reads in
   ``read_only`` streaming mode so we never hold a parsed workbook in memory.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook

from .naming import dedupe_labels, tidy_label

MAX_HEADER_SCAN_ROWS = 25
"""How many leading rows to inspect when hunting for the header."""


@dataclass
class SheetInfo:
    """What we know about one sheet before committing to a full read."""

    name: str
    header_row: int  # 1-based, as shown in Excel
    columns: List[str] = field(default_factory=list)
    n_columns: int = 0

    def __str__(self) -> str:  # pragma: no cover - display helper
        return f"{self.name} (header row {self.header_row}, {self.n_columns} cols)"


def list_sheets(path_or_buffer) -> List[str]:
    """Return sheet names without parsing any cell data."""
    workbook = load_workbook(path_or_buffer, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _score_row(values: Sequence[object]) -> float:
    """Score how header-like a row looks.

    A header row is mostly short, distinct, non-numeric text.  Data rows carry
    numbers and repeat values; title rows have one cell and lots of blanks.
    """
    cells = [tidy_label(v) for v in values]
    filled = [c for c in cells if c]
    if len(filled) < 3:
        return 0.0

    texty = [c for c in filled if not _looks_numeric(c)]
    distinct = len(set(filled)) / len(filled)
    density = len(filled) / max(len(cells), 1)
    text_share = len(texty) / len(filled)
    # Headers are short labels; penalise essay-length cells.
    brevity = sum(1 for c in filled if len(c) <= 45) / len(filled)

    return (len(filled) ** 0.5) * distinct * density * text_share * brevity


def _looks_numeric(text: str) -> bool:
    try:
        float(text.replace(",", "").replace("$", ""))
        return True
    except ValueError:
        return False


def detect_header_row(path_or_buffer, sheet_name: str) -> SheetInfo:
    """Find the header row of ``sheet_name`` and read its labels."""
    workbook = load_workbook(path_or_buffer, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        best_row, best_score, best_values = 1, -1.0, []
        for row_index, row in enumerate(
            worksheet.iter_rows(max_row=MAX_HEADER_SCAN_ROWS, values_only=True),
            start=1,
        ):
            score = _score_row(row)
            if score > best_score:
                best_row, best_score, best_values = row_index, score, list(row)

        # Trim trailing all-blank columns before naming.
        while best_values and tidy_label(best_values[-1]) == "":
            best_values.pop()

        columns = dedupe_labels(best_values)
        return SheetInfo(
            name=sheet_name,
            header_row=best_row,
            columns=columns,
            n_columns=len(columns),
        )
    finally:
        workbook.close()


def profile_workbook(path_or_buffer) -> Dict[str, SheetInfo]:
    """Detect the header row for every sheet in a workbook."""
    return {
        name: detect_header_row(path_or_buffer, name)
        for name in list_sheets(path_or_buffer)
    }


def read_sheet(
    path_or_buffer,
    sheet_name: str,
    header_row: Optional[int] = None,
    max_rows: Optional[int] = None,
) -> pd.DataFrame:
    """Read one sheet into a DataFrame.

    ``header_row`` is 1-based (what Excel shows).  When omitted it is sniffed.
    Formula cells resolve to their cached values, so ``=TRIM(C3)`` arrives as
    the trimmed text rather than the formula -- unless the workbook was written
    by a tool that never cached results, in which case the cell reads as blank.
    """
    if header_row is None:
        header_row = detect_header_row(path_or_buffer, sheet_name).header_row

    frame = pd.read_excel(
        path_or_buffer,
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
        nrows=max_rows,
    )
    frame.columns = dedupe_labels(frame.columns)
    # Drop columns that are entirely empty and carry a placeholder name.
    empty_placeholders = [
        c
        for c in frame.columns
        if c.startswith("Column_") and frame[c].isna().all()
    ]
    return frame.drop(columns=empty_placeholders)


def formula_leak_ratio(frame: pd.DataFrame) -> float:
    """Share of cells that still contain a raw ``=FORMULA`` string.

    A non-zero value means the source workbook had no cached formula results,
    so those columns need recalculating in Excel before the data is usable.
    """
    total = 0
    leaked = 0
    for column in frame.columns:
        series = frame[column]
        if series.dtype == object or str(series.dtype) == "str":
            text = series.dropna().astype(str)
            total += len(text)
            leaked += int(text.str.startswith("=").sum())
    return (leaked / total) if total else 0.0
